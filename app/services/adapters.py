"""Format adapters: every source format becomes a CanonicalDocument.

Three adapters, one output shape:

    Recordizer    CSV, TSV, XLSX, JSON, JSONL -> one RECORD block per row, with
                  fields preserved so entities come from columns, not from NER.
    ProseParser   TXT, MD, RST -> HEADING and PROSE blocks following the markdown
                  hierarchy, plus TABLE blocks for markdown tables.
    RichAdapter   PDF, DOCX, PPTX, HTML via Docling -> markdown with table
                  structure intact, then routed through ProseParser.

Format handling stops here. Nothing downstream needs to know where a document
came from, only what kind of blocks it contains.
"""
from __future__ import annotations

import csv
import json
import logging
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from app.models.canonical import (
    BlockKind,
    CanonicalDocument,
    ContentBlock,
    Provenance,
    TableCell,
)

logger = logging.getLogger(__name__)

STRUCTURED_SUFFIXES = {".csv", ".tsv", ".xlsx", ".xls", ".json", ".jsonl"}
PROSE_SUFFIXES = {".txt", ".md", ".markdown", ".rst"}
RICH_SUFFIXES = {".pdf", ".docx", ".pptx", ".html", ".htm"}

_HEADING_RE = re.compile(r"^(#{1,6})\s+(.+?)\s*#*$")
_MD_TABLE_SEP = re.compile(r"^\s*\|?[\s:|-]+\|[\s:|-]*$")
_NULL_VALUES = {
    "", "na", "n/a", "none", "null", "nil", "-", "--", "unknown",
    "not applicable", "not specified", "nan",
}


def _slug(value: str, limit: int = 60) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9]+", "_", str(value)).strip("_").lower()
    return cleaned[:limit] or "untitled"


def _is_null(value: Any) -> bool:
    return value is None or str(value).strip().lower() in _NULL_VALUES


# --------------------------------------------------------------------- structured
class Recordizer:
    """Turns tabular sources into one RECORD block per row.

    Fields are kept alongside the rendered text. A row's entities are already
    separated into columns, so preserving that structure lets extraction read them
    rather than infer them — the difference between confidence 1.0 and 0.8.
    """

    def load(
        self, path: Path, max_rows: Optional[int] = None
    ) -> Tuple[List[Dict[str, Any]], List[str]]:
        suffix = path.suffix.lower()
        if suffix in (".csv", ".tsv"):
            return self._load_delimited(path, max_rows, "\t" if suffix == ".tsv" else ",")
        if suffix in (".xlsx", ".xls"):
            return self._load_excel(path, max_rows)
        if suffix == ".jsonl":
            return self._load_jsonl(path, max_rows)
        if suffix == ".json":
            return self._load_json(path, max_rows)
        raise ValueError(f"Recordizer cannot read {suffix}")

    def _load_delimited(
        self, path: Path, max_rows: Optional[int], delimiter: str
    ) -> Tuple[List[Dict[str, Any]], List[str]]:
        # Text columns in real datasets routinely exceed the default field limit.
        csv.field_size_limit(min(sys.maxsize, 2_147_483_647))
        rows: List[Dict[str, Any]] = []
        with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
            reader = csv.DictReader(handle, delimiter=delimiter)
            columns = list(reader.fieldnames or [])
            for index, row in enumerate(reader):
                if max_rows and index >= max_rows:
                    break
                rows.append(row)
        return rows, columns

    def _load_excel(
        self, path: Path, max_rows: Optional[int]
    ) -> Tuple[List[Dict[str, Any]], List[str]]:
        try:
            from openpyxl import load_workbook  # noqa: PLC0415
        except ImportError as exc:
            raise RuntimeError(
                "Reading .xlsx requires openpyxl, or export the sheet to .csv."
            ) from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in next(iterator, [])]
        rows: List[Dict[str, Any]] = []
        for index, values in enumerate(iterator):
            if max_rows and index >= max_rows:
                break
            rows.append({h: ("" if v is None else str(v)) for h, v in zip(header, values)})
        workbook.close()
        return rows, header

    def _load_jsonl(
        self, path: Path, max_rows: Optional[int]
    ) -> Tuple[List[Dict[str, Any]], List[str]]:
        rows: List[Dict[str, Any]] = []
        columns: List[str] = []
        with path.open("r", encoding="utf-8", errors="replace") as handle:
            for index, line in enumerate(handle):
                if max_rows and index >= max_rows:
                    break
                line = line.strip()
                if not line:
                    continue
                record = json.loads(line)
                if isinstance(record, dict):
                    rows.append(record)
                    columns.extend(k for k in record if k not in columns)
        return rows, columns

    def _load_json(
        self, path: Path, max_rows: Optional[int]
    ) -> Tuple[List[Dict[str, Any]], List[str]]:
        data = json.loads(path.read_text(encoding="utf-8", errors="replace"))
        if isinstance(data, dict):
            # Accept {"records": [...]} and similar wrappers.
            for value in data.values():
                if isinstance(value, list) and value and isinstance(value[0], dict):
                    data = value
                    break
            else:
                data = [data]
        rows = [r for r in data if isinstance(r, dict)][: max_rows or len(data)]
        columns: List[str] = []
        for row in rows:
            columns.extend(k for k in row if k not in columns)
        return rows, columns

    def to_document(
        self,
        path: Path,
        rows: Sequence[Dict[str, Any]],
        columns: Sequence[str],
        subject_column: Optional[str] = None,
    ) -> CanonicalDocument:
        """Build a CanonicalDocument with one RECORD block per row."""
        doc_provenance = Provenance(
            source_uri=str(path),
            source_format=path.suffix.lstrip("."),
            extractor="recordizer",
        )
        blocks: List[ContentBlock] = []

        for index, row in enumerate(rows):
            fields = [
                TableCell(column=str(col), value=str(row.get(col, "")).strip())
                for col in columns
                if not _is_null(row.get(col))
            ]
            if not fields:
                continue

            subject = (
                str(row.get(subject_column, "")).strip() if subject_column else ""
            )
            text = self._render_record(fields, subject)
            blocks.append(
                ContentBlock(
                    block_id=f"{_slug(path.stem)}_r{index}",
                    kind=BlockKind.RECORD,
                    text=text,
                    fields=fields,
                    table_headers=list(columns),
                    provenance=Provenance(
                        source_uri=str(path),
                        source_format=path.suffix.lstrip("."),
                        row_index=index,
                        extractor="recordizer",
                    ),
                    metadata={"subject": subject} if subject else {},
                )
            )

        return CanonicalDocument(
            doc_id=f"doc_{_slug(path.stem)}",
            title=path.stem,
            blocks=blocks,
            provenance=doc_provenance,
            metadata={"rows": len(rows), "columns": list(columns)},
        )

    @staticmethod
    def _render_record(fields: Sequence[TableCell], subject: str) -> str:
        """Render a row as labelled lines.

        Vector search matches text against text, and a raw comma-separated row
        gives the embedding model no grammar to work with. Each value is rendered
        beside its column name so the chunk is self-describing.
        """
        lines = [f"{cell.column}: {cell.value}" for cell in fields]
        body = "\n".join(lines)
        return f"# {subject}\n\n{body}" if subject else body


# ------------------------------------------------------------------------- prose
class ProseParser:
    """Parses markdown or plain text into HEADING, PROSE, TABLE, and LIST blocks."""

    def to_document(
        self, path: Path, text: Optional[str] = None, source_format: Optional[str] = None
    ) -> CanonicalDocument:
        content = text if text is not None else path.read_text(encoding="utf-8", errors="replace")
        fmt = source_format or path.suffix.lstrip(".")
        doc_provenance = Provenance(
            source_uri=str(path), source_format=fmt, extractor="prose_parser"
        )

        blocks = self._parse_blocks(content, path, fmt)
        title = next(
            (b.text for b in blocks if b.kind is BlockKind.HEADING), path.stem
        )

        return CanonicalDocument(
            doc_id=f"doc_{_slug(path.stem)}",
            title=title,
            blocks=blocks,
            provenance=doc_provenance,
        )

    def _parse_blocks(self, content: str, path: Path, fmt: str) -> List[ContentBlock]:
        blocks: List[ContentBlock] = []
        heading_stack: List[Tuple[int, str]] = []
        buffer: List[str] = []
        offset = 0
        counter = 0

        def flush_prose(end_offset: int) -> None:
            nonlocal buffer, counter
            body = "\n".join(buffer).strip()
            buffer = []
            if not body:
                return
            for para in re.split(r"\n\s*\n", body):
                para = para.strip()
                if not para:
                    continue
                counter += 1
                kind = (
                    BlockKind.LIST
                    if re.match(r"^\s*[-*+]\s+|^\s*\d+[.)]\s+", para)
                    else BlockKind.PROSE
                )
                blocks.append(
                    ContentBlock(
                        block_id=f"{_slug(path.stem)}_b{counter}",
                        kind=kind,
                        text=para,
                        provenance=Provenance(
                            source_uri=str(path),
                            source_format=fmt,
                            section_path=[h[1] for h in heading_stack],
                            char_end=end_offset,
                            extractor="prose_parser",
                        ),
                    )
                )

        lines = content.splitlines()
        index = 0
        while index < len(lines):
            line = lines[index]
            offset += len(line) + 1

            heading = _HEADING_RE.match(line.strip())
            if heading:
                flush_prose(offset)
                level = len(heading.group(1))
                title = heading.group(2).strip()
                while heading_stack and heading_stack[-1][0] >= level:
                    heading_stack.pop()
                heading_stack.append((level, title))
                counter += 1
                blocks.append(
                    ContentBlock(
                        block_id=f"{_slug(path.stem)}_b{counter}",
                        kind=BlockKind.HEADING,
                        text=title,
                        heading_level=level,
                        provenance=Provenance(
                            source_uri=str(path),
                            source_format=fmt,
                            section_path=[h[1] for h in heading_stack[:-1]],
                            extractor="prose_parser",
                        ),
                    )
                )
                index += 1
                continue

            # A markdown table: a pipe row followed by a separator row.
            if "|" in line and index + 1 < len(lines) and _MD_TABLE_SEP.match(lines[index + 1]):
                flush_prose(offset)
                table_lines: List[str] = []
                while index < len(lines) and "|" in lines[index]:
                    table_lines.append(lines[index])
                    index += 1
                counter += 1
                block = self._table_block(
                    table_lines, f"{_slug(path.stem)}_b{counter}", path, fmt,
                    [h[1] for h in heading_stack],
                )
                if block:
                    blocks.append(block)
                continue

            buffer.append(line)
            index += 1

        flush_prose(offset)
        return blocks

    @staticmethod
    def _table_block(
        lines: List[str], block_id: str, path: Path, fmt: str, section_path: List[str]
    ) -> Optional[ContentBlock]:
        """Parse a markdown table, keeping headers and rows separate from the text."""
        def cells(row: str) -> List[str]:
            return [c.strip() for c in row.strip().strip("|").split("|")]

        if len(lines) < 2:
            return None
        headers = cells(lines[0])
        rows = [cells(line) for line in lines[2:] if line.strip()]
        if not headers:
            return None

        return ContentBlock(
            block_id=block_id,
            kind=BlockKind.TABLE,
            text="\n".join(lines),
            table_headers=headers,
            table_rows=rows,
            provenance=Provenance(
                source_uri=str(path),
                source_format=fmt,
                section_path=section_path,
                extractor="prose_parser",
            ),
        )


# -------------------------------------------------------------------------- rich
class RichAdapter:
    """PDF, DOCX, PPTX, HTML via Docling, preserving table structure."""

    def __init__(self) -> None:
        self._converter = None

    def _get_converter(self):
        if self._converter is None:
            try:
                from docling.document_converter import DocumentConverter  # noqa: PLC0415
            except ImportError as exc:
                raise RuntimeError(
                    "Reading PDF/DOCX/PPTX/HTML requires Docling: pip install docling\n"
                    "Alternatively convert the file to .txt, .md, or .csv first."
                ) from exc
            self._converter = DocumentConverter()
        return self._converter

    def to_document(self, path: Path, prose_parser: ProseParser) -> CanonicalDocument:
        """Convert to markdown, then reuse the prose parser for block structure.

        Docling preserves tables as markdown tables, so the prose parser's table
        handling recovers them as TABLE blocks rather than flattening them.
        """
        result = self._get_converter().convert(str(path))
        markdown = result.document.export_to_markdown()
        if not markdown.strip():
            return CanonicalDocument(
                doc_id=f"doc_{_slug(path.stem)}",
                title=path.stem,
                provenance=Provenance(
                    source_uri=str(path),
                    source_format=path.suffix.lstrip("."),
                    extractor="docling",
                ),
            )

        document = prose_parser.to_document(
            path, text=markdown, source_format=path.suffix.lstrip(".")
        )
        for block in document.blocks:
            block.provenance.extractor = "docling"
        document.provenance.extractor = "docling"
        return document


# ------------------------------------------------------------------ orchestration
class AdapterRegistry:
    """Detects a file's format and routes it to the right adapter."""

    def __init__(self) -> None:
        self.recordizer = Recordizer()
        self.prose_parser = ProseParser()
        self.rich_adapter = RichAdapter()

    @staticmethod
    def detect(path: Path) -> str:
        suffix = path.suffix.lower()
        if suffix in STRUCTURED_SUFFIXES:
            return "structured"
        if suffix in PROSE_SUFFIXES:
            return "prose"
        if suffix in RICH_SUFFIXES:
            return "rich"
        return "unsupported"

    def to_canonical(
        self,
        path: Path,
        max_rows: Optional[int] = None,
        subject_column: Optional[str] = None,
    ) -> CanonicalDocument:
        """Convert any supported file into a CanonicalDocument."""
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"Source not found: {path}")

        kind = self.detect(path)

        if kind == "structured":
            rows, columns = self.recordizer.load(path, max_rows)
            return self.recordizer.to_document(path, rows, columns, subject_column)

        if kind == "prose":
            return self.prose_parser.to_document(path)

        if kind == "rich":
            return self.rich_adapter.to_document(path, self.prose_parser)

        raise ValueError(
            f"Unsupported format '{path.suffix}'. Supported: "
            f"{sorted(STRUCTURED_SUFFIXES | PROSE_SUFFIXES | RICH_SUFFIXES)}"
        )


adapter_registry = AdapterRegistry()
