"""Structured-source ingestion: CSV, XLSX, JSON, and JSONL into a graph.

Tabular data does not need the prose pipeline. The entities are already separated
into columns, so running NER over a cell is strictly worse than reading it: the
column header states the type, the cell states the value, and co-occurrence in a
row states the relation. Extraction confidence is 1.0 rather than a heuristic
score, because nothing is being inferred.

The pipeline:

    profile columns   -> classify each as entity / attribute / freetext / ignore
    build entities    -> one canonical node per distinct value
    build relations   -> co-occurring entity columns become edges
    verbalize rows    -> a natural-language sentence per row, for embedding

Classification is by cardinality, which is a reliable signal in practice:
a column repeating a small set of values across many rows is categorical;
one with many distinct short values is an entity; one with long unique values is
free text that belongs in the chunk, not the graph.
"""
from __future__ import annotations

import csv
import json
import logging
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

from app.core.tenant_schema import TenantGraphSchema
from app.models.graph import Edge, Vertex
from app.services.extraction_service import entity_id_for, normalize_entity_name

logger = logging.getLogger(__name__)

# Cells frequently hold several values: "Tulsi, Ashwagandha" or "Vata; Kapha".
_MULTI_VALUE_SPLIT = re.compile(r"\s*[;,/|]\s*|\s+and\s+", re.IGNORECASE)

# Values that carry no information and should never become nodes.
_NULL_VALUES = {
    "", "na", "n/a", "none", "null", "nil", "-", "--", "unknown", "not applicable",
    "not specified", "varies", "any", "all", "other", "others", "nan",
}

# Column-name hints that override cardinality-based classification.
# Order matters: the first match wins, so the narrowest hints come first.
_NAME_HINTS: List[Tuple[str, str]] = [
    (r"\b(?:id|uuid|index|key|row|serial|no|number)\b", "ignore"),
    (r"\b(?:url|link|href|image|photo|src)\b", "ignore"),
    # Free text before entity: "Patient Recommendations" contains "patient" but is
    # advice prose, not a person, and "Diet and Lifestyle Recommendations" is not a
    # named entity either.
    (r"\b(?:description|summary|notes?|comment|text|content|detail|recommendation|"
     r"recommendations|instruction|advice|prognosis|diagnosis|guideline)\b", "freetext"),
    # Categorical before entity: "Symptom Severity" contains "symptom" but is a
    # graded value, and "Duration of Treatment" contains "treatment" but is a span.
    (r"\b(?:severity|duration|frequency|type|category|class|group|status|level|"
     r"stage|gender|sex|season|seasonal|dosha|doshas|prakriti|constitution|"
     r"age|grade|rating|score|region|pattern|patterns)\b", "attribute"),
    (r"\b(?:name|title|disease|condition|herb|herbs|drug|product|company|"
     r"organization|person|author|entity|item|ingredient|symptom|symptoms|"
     r"treatment|remedy|remedies|formulation|therapy)\b", "entity"),
]

# Columns holding a transliteration or translation of another column. These are
# aliases of an existing entity, not new entities, and promoting them creates a
# duplicate node per language.
_ALIAS_COLUMN = re.compile(
    r"\b(?:hindi|marathi|tamil|telugu|bengali|sanskrit|urdu|arabic|chinese|"
    r"spanish|french|german|latin|local|native|vernacular|regional|alternate|"
    r"alternative|other)\s+(?:name|title|term)\b|\b(?:name|title)\s+in\s+\w+\b",
    re.IGNORECASE,
)


@dataclass
class ColumnProfile:
    """What a column contains and how it should map into the graph."""

    name: str
    role: str                      # entity | attribute | freetext | ignore
    distinct: int
    total: int
    mean_length: float
    multi_valued: bool
    label: str = "Entity"          # graph label for entity/attribute roles
    samples: List[str] = field(default_factory=list)

    @property
    def cardinality_ratio(self) -> float:
        return self.distinct / self.total if self.total else 0.0

    def to_dict(self) -> Dict[str, Any]:
        return {
            "name": self.name,
            "role": self.role,
            "label": self.label,
            "distinct": self.distinct,
            "total": self.total,
            "cardinality_ratio": round(self.cardinality_ratio, 3),
            "mean_length": round(self.mean_length, 1),
            "multi_valued": self.multi_valued,
            "samples": self.samples[:3],
        }


@dataclass
class StructuredDocument:
    """One source row, verbalized for embedding and linked to its entities."""

    doc_id: str
    text: str
    entities: List[Vertex]
    edges: List[Edge]
    metadata: Dict[str, Any] = field(default_factory=dict)


def split_values(raw: str) -> List[str]:
    """Split a possibly multi-valued cell into individual values."""
    if not raw:
        return []
    parts = [p.strip(" .;,") for p in _MULTI_VALUE_SPLIT.split(raw)]
    return [p for p in parts if p and p.lower() not in _NULL_VALUES]


def is_null(value: Any) -> bool:
    return value is None or str(value).strip().lower() in _NULL_VALUES


class ColumnProfiler:
    """Classifies columns by name hints and value cardinality."""

    # Tuned against real datasets: a column repeating <=40 values across many rows
    # is categorical; beyond that, short distinct values are entities.
    MAX_ATTRIBUTE_DISTINCT = 40
    MAX_ATTRIBUTE_LENGTH = 40
    MAX_ENTITY_LENGTH = 80
    MIN_ROWS_FOR_RATIO = 20

    def profile(self, rows: Sequence[Dict[str, Any]], columns: Sequence[str]) -> List[ColumnProfile]:
        profiles: List[ColumnProfile] = []
        total = len(rows)

        for column in columns:
            raw_values = [str(r.get(column, "")).strip() for r in rows]
            present = [v for v in raw_values if not is_null(v)]
            if not present:
                profiles.append(
                    ColumnProfile(column, "ignore", 0, total, 0.0, False, samples=[])
                )
                continue

            multi = any(_MULTI_VALUE_SPLIT.search(v) for v in present[:200])
            expanded: List[str] = []
            for value in present:
                expanded.extend(split_values(value) if multi else [value])

            distinct_values = {v.lower() for v in expanded}
            mean_length = sum(len(v) for v in expanded) / len(expanded)

            role, label = self._classify(
                column, len(distinct_values), len(expanded), mean_length
            )
            profiles.append(
                ColumnProfile(
                    name=column,
                    role=role,
                    distinct=len(distinct_values),
                    total=total,
                    mean_length=mean_length,
                    multi_valued=multi,
                    label=label,
                    samples=list(dict.fromkeys(expanded))[:5],
                )
            )
        return profiles

    def _classify(
        self, column: str, distinct: int, occurrences: int, mean_length: float
    ) -> Tuple[str, str]:
        """Return (role, graph_label) for a column."""
        lowered = column.lower()

        # A translated name is an alias of an entity that already exists in another
        # column; promoting it would create a duplicate node per language.
        if _ALIAS_COLUMN.search(lowered):
            return "alias", "Entity"

        # Name hints are more reliable than statistics when they match.
        for pattern, hinted_role in _NAME_HINTS:
            if re.search(pattern, lowered):
                if hinted_role == "ignore":
                    return "ignore", "Entity"
                if hinted_role == "freetext":
                    return "freetext", "Entity"
                if hinted_role == "attribute":
                    return "attribute", "Attribute"
                # An "entity"-hinted column that is actually long prose is prose.
                if mean_length > self.MAX_ENTITY_LENGTH:
                    return "freetext", "Entity"
                return "entity", self._entity_label(lowered)

        if mean_length > self.MAX_ENTITY_LENGTH:
            return "freetext", "Entity"

        # A near-unique column is an identifier, not a shared entity: making it a
        # node produces one orphan per row and no traversable structure.
        if occurrences >= self.MIN_ROWS_FOR_RATIO and distinct / occurrences > 0.95:
            return "ignore", "Entity"

        if distinct <= self.MAX_ATTRIBUTE_DISTINCT and mean_length <= self.MAX_ATTRIBUTE_LENGTH:
            return "attribute", "Attribute"

        return "entity", self._entity_label(lowered)

    @staticmethod
    def _entity_label(column_lower: str) -> str:
        """Map a column name onto a generic vertex label."""
        if re.search(r"\b(?:person|author|doctor|patient|practitioner|name)\b", column_lower):
            return "Person"
        if re.search(r"\b(?:company|organization|institution|manufacturer|brand)\b", column_lower):
            return "Organization"
        if re.search(r"\b(?:product|formulation|remedy|medicine|drug|item|artifact|"
                     r"document|book|text)\b", column_lower):
            return "Artifact"
        if re.search(r"\b(?:disease|condition|symptom|concept|topic|technique|"
                     r"method|therapy|treatment|practice)\b", column_lower):
            return "Concept"
        return "Entity"


class StructuredIngestionService:
    """Converts tabular sources into graph entities, relations, and text."""

    def __init__(self) -> None:
        self.profiler = ColumnProfiler()

    # ------------------------------------------------------------------ loading
    def load_rows(self, path: Path, max_rows: Optional[int] = None) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Read rows from CSV, XLSX, JSON, or JSONL."""
        suffix = path.suffix.lower()

        if suffix in (".csv", ".tsv"):
            return self._load_csv(path, max_rows, delimiter="\t" if suffix == ".tsv" else ",")
        if suffix in (".xlsx", ".xls"):
            return self._load_excel(path, max_rows)
        if suffix == ".jsonl":
            return self._load_jsonl(path, max_rows)
        if suffix == ".json":
            return self._load_json(path, max_rows)
        raise ValueError(f"Unsupported structured format: {suffix}")

    def _load_csv(self, path: Path, max_rows: Optional[int], delimiter: str) -> Tuple[List[Dict[str, Any]], List[str]]:
        # Text columns in real datasets routinely exceed the default field limit.
        csv.field_size_limit(min(sys.maxsize, 2_147_483_647))
        rows: List[Dict[str, Any]] = []
        with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
            reader = csv.DictReader(handle, delimiter=delimiter)
            columns = list(reader.fieldnames or [])
            for index, row in enumerate(reader):
                if max_rows and index >= max_rows:
                    break
                rows.append(row)
        return rows, columns

    def _load_excel(self, path: Path, max_rows: Optional[int]) -> Tuple[List[Dict[str, Any]], List[str]]:
        try:
            from openpyxl import load_workbook  # noqa: PLC0415
        except ImportError as exc:
            raise RuntimeError(
                "Reading .xlsx requires openpyxl. Install it, or use the .csv export."
            ) from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in next(iterator, [])]
        rows: List[Dict[str, Any]] = []
        for index, values in enumerate(iterator):
            if max_rows and index >= max_rows:
                break
            rows.append({h: ("" if v is None else str(v)) for h, v in zip(header, values)})
        workbook.close()
        return rows, header

    def _load_jsonl(self, path: Path, max_rows: Optional[int]) -> Tuple[List[Dict[str, Any]], List[str]]:
        rows: List[Dict[str, Any]] = []
        columns: List[str] = []
        with path.open("r", encoding="utf-8", errors="replace") as handle:
            for index, line in enumerate(handle):
                if max_rows and index >= max_rows:
                    break
                line = line.strip()
                if not line:
                    continue
                record = json.loads(line)
                if isinstance(record, dict):
                    rows.append(record)
                    for key in record:
                        if key not in columns:
                            columns.append(key)
        return rows, columns

    def _load_json(self, path: Path, max_rows: Optional[int]) -> Tuple[List[Dict[str, Any]], List[str]]:
        data = json.loads(path.read_text(encoding="utf-8", errors="replace"))
        if isinstance(data, dict):
            # Accept {"records": [...]} and similar wrappers.
            for value in data.values():
                if isinstance(value, list) and value and isinstance(value[0], dict):
                    data = value
                    break
            else:
                data = [data]
        rows = [r for r in data if isinstance(r, dict)][: max_rows or len(data)]
        columns: List[str] = []
        for row in rows:
            for key in row:
                if key not in columns:
                    columns.append(key)
        return rows, columns

    # ------------------------------------------------------------------ mapping
    def build_documents(
        self,
        rows: Sequence[Dict[str, Any]],
        profiles: Sequence[ColumnProfile],
        schema: TenantGraphSchema,
        doc_prefix: str = "row",
    ) -> List[StructuredDocument]:
        """Convert profiled rows into verbalized documents with entities and edges."""
        entity_cols = [p for p in profiles if p.role == "entity"]
        attribute_cols = [p for p in profiles if p.role == "attribute"]
        alias_cols = [p for p in profiles if p.role == "alias"]

        subject_col = self._pick_subject(entity_cols, profiles)

        documents: List[StructuredDocument] = []

        for index, row in enumerate(rows):
            vertices: Dict[str, Vertex] = {}
            edges: List[Edge] = []

            def add_vertex(value: str, label: str) -> Optional[str]:
                normalized = normalize_entity_name(value)
                if not normalized or not schema.validate_vertex_label(label):
                    return None
                entity_id = entity_id_for(value, label)
                if entity_id not in vertices:
                    vertices[entity_id] = Vertex(
                        id=entity_id,
                        label=label,
                        properties={
                            "name": value,
                            "normalized_name": normalized,
                            "entity_id": entity_id,
                            "confidence": 1.0,
                            "extractor": "structured",
                        },
                    )
                return entity_id

            # Subject entities for this row.
            subject_ids: List[str] = []
            if subject_col:
                for value in self._cell_values(row, subject_col):
                    entity_id = add_vertex(value, subject_col.label)
                    if entity_id:
                        subject_ids.append(entity_id)

            # Translated names attach to the subject as aliases, so a query in any
            # of those languages links to the same canonical node.
            if subject_ids and alias_cols:
                aliases: List[str] = []
                for column in alias_cols:
                    aliases.extend(self._cell_values(row, column))
                if aliases:
                    for subject_id in subject_ids:
                        existing = vertices[subject_id].properties.setdefault("aliases", [])
                        for alias in aliases:
                            if alias not in existing:
                                existing.append(alias)

            # Other entity columns relate to the subject. Co-occurrence in a row is
            # the relation, stated rather than inferred, so confidence is 1.0.
            for column in entity_cols:
                if subject_col and column.name == subject_col.name:
                    continue
                for value in self._cell_values(row, column):
                    entity_id = add_vertex(value, column.label)
                    if not entity_id:
                        continue
                    edge_type = self._infer_edge_type(column.name, schema)
                    for subject_id in subject_ids:
                        edges.append(
                            Edge(
                                source=subject_id,
                                target=entity_id,
                                type=edge_type,
                                properties={
                                    "confidence": 1.0,
                                    "source_column": column.name,
                                    "provenance": "structured",
                                },
                            )
                        )

            # Categorical columns become HAS_ATTRIBUTE edges.
            for column in attribute_cols:
                for value in self._cell_values(row, column):
                    entity_id = add_vertex(value, "Attribute")
                    if not entity_id:
                        continue
                    for subject_id in subject_ids:
                        edges.append(
                            Edge(
                                source=subject_id,
                                target=entity_id,
                                type="HAS_ATTRIBUTE",
                                properties={
                                    "confidence": 1.0,
                                    "source_column": column.name,
                                    "provenance": "structured",
                                },
                            )
                        )

            text = self.verbalize_row(row, profiles)
            if not text.strip():
                continue

            doc_id = self._row_id(row, subject_col, index, doc_prefix)
            documents.append(
                StructuredDocument(
                    doc_id=doc_id,
                    text=text,
                    entities=list(vertices.values()),
                    edges=edges,
                    metadata={
                        "source": "structured",
                        "row_index": index,
                        "subject": row.get(subject_col.name, "") if subject_col else "",
                    },
                )
            )

        return documents

    @staticmethod
    def _pick_subject(
        entity_cols: Sequence[ColumnProfile], profiles: Sequence[ColumnProfile]
    ) -> Optional[ColumnProfile]:
        """Choose the column each row is *about*.

        This anchors every relation in the row, so getting it wrong inverts the
        whole graph. Three signals, strongest first:

        1. A column that an alias column translates ("Hindi Name" implies the
           subject is whatever "Name"/"Disease" column it mirrors) - the dataset
           itself is telling us what the row is about.
        2. Position: tabular data conventionally leads with its subject.
        3. Single-valued over multi-valued: a row is about one thing, and a cell
           listing several values is a property of the subject, not the subject.

        Cardinality alone is unreliable: a small controlled vocabulary (36 herbs)
        looks "most repeated" while being an attribute of the row, not its topic.
        """
        if not entity_cols:
            return None

        has_alias = any(p.role == "alias" for p in profiles)
        order = {p.name: i for i, p in enumerate(profiles)}

        def score(column: ColumnProfile) -> tuple:
            first_entity = order.get(column.name, 99) == min(
                order.get(c.name, 99) for c in entity_cols
            )
            return (
                # An alias column present means the leading entity column is the
                # translated subject.
                has_alias and first_entity,
                not column.multi_valued,
                -order.get(column.name, 99),
            )

        return max(entity_cols, key=score)

    @staticmethod
    def _cell_values(row: Dict[str, Any], column: ColumnProfile) -> List[str]:
        raw = str(row.get(column.name, "") or "").strip()
        if is_null(raw):
            return []
        return split_values(raw) if column.multi_valued else [raw]

    @staticmethod
    def _infer_edge_type(column_name: str, schema: TenantGraphSchema) -> str:
        """Choose the generic edge type that best fits a column's meaning."""
        lowered = column_name.lower()
        candidates: List[Tuple[str, str]] = [
            (r"\b(?:symptom|effect|complication|risk|cause|treat|remedy|herb|"
             r"medicine|drug|intervention|therapy|prevention)\b", "AFFECTS"),
            (r"\b(?:ingredient|component|part|contains|composition)\b", "PART_OF"),
            (r"\b(?:source|origin|derived|extracted|from)\b", "DERIVED_FROM"),
            (r"\b(?:reference|citation|cited|source_doc)\b", "CITES"),
            (r"\b(?:history|prior|previous|before|after|next|stage)\b", "PRECEDES"),
        ]
        for pattern, edge_type in candidates:
            if re.search(pattern, lowered) and schema.validate_edge_type(edge_type):
                return edge_type
        return "ASSOCIATED_WITH" if schema.validate_edge_type("ASSOCIATED_WITH") else "RELATED_TO"

    @staticmethod
    def _row_id(
        row: Dict[str, Any], subject_col: Optional[ColumnProfile], index: int, prefix: str
    ) -> str:
        if subject_col:
            subject = normalize_entity_name(str(row.get(subject_col.name, "")))
            if subject:
                return f"{prefix}_{subject[:50]}_{index}"
        return f"{prefix}_{index}"

    @staticmethod
    def verbalize_row(row: Dict[str, Any], profiles: Sequence[ColumnProfile]) -> str:
        """Turn a row into natural language for embedding.

        Vector search matches text against text. Embedding a raw comma-separated
        row gives the model no grammar to work with, so each field becomes a
        labelled sentence instead.
        """
        lines: List[str] = []
        subject_line: Optional[str] = None

        for profile in profiles:
            if profile.role == "ignore":
                continue
            raw = str(row.get(profile.name, "") or "").strip()
            if is_null(raw):
                continue
            label = profile.name.strip()
            if subject_line is None and profile.role == "entity":
                subject_line = f"# {raw}"
                lines.append(f"{label}: {raw}")
                continue
            lines.append(f"{label}: {raw}")

        body = "\n".join(lines)
        return f"{subject_line}\n\n{body}" if subject_line else body

    # ------------------------------------------------------------------ public
    def analyze(
        self, path: Path, max_rows: Optional[int] = None
    ) -> Tuple[List[Dict[str, Any]], List[str], List[ColumnProfile]]:
        """Load and profile a structured source without ingesting it."""
        rows, columns = self.load_rows(path, max_rows)
        profiles = self.profiler.profile(rows, columns)
        return rows, columns, profiles


structured_ingestion_service = StructuredIngestionService()
